#! python3
# Get input from excel file -> Find email -> Directly update excel file

import sys
import requests
import re
import openpyxl
import os
from bs4 import BeautifulSoup
from googlesearch import search
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, Text, messagebox
import threading

# TODO: handle urllib.error.HTTPError: HTTP Error 429: Too Many Requests
# TODO: handle time in a web
# TODO: get "at" email with re
# TODO: handle Email Address Obfuscation
# TODO: handle "farinaz (at) ucsd (dot) edu"/"girard at lirmm dot fr" email
# TODO: exponential backoff
# TODO: Stop button


def is_valid_email(email):
    invalid = [
        'admin', 'admissions', 'app', 'ask', 'biochem', 'business', 
        'career', 'college', 'communications', 'contact', 'contact', 'copy', "council",
        'director', 'engineer', 'enquiries', 'experts', 'help', 
        'idm', 'idm', 'info', 'inquiry', 'international', 'job', 'kontakt', 
        'lab', 'life', 'news', 'medcentral', "medicine", 'office', "order",
        'phd', 'president', 'press', 'profile', 'profiles', 'program', "physic", 
        'registrator', 'reply', 'research', 'researcher', 
        'sales', 'service', 'staff', 'support', "student",
        'team', 'web', "xxx"]
        
    first, _ = email.split('@')
    for invalid_word in invalid:
        if invalid_word in first.lower():
            return False
    return True


def extract_mailto(soup):
    emails = []
    mailtos = soup.select('a[href^="mailto"]')
    if len(mailtos) == 0:
        mailtos = soup.select('a[href^="Mailto"]')

    for i in mailtos:
        href = i['href']
        try:
            _, mail = href.split(':')
        except ValueError as e:
            print(e)
            continue

        if not is_valid_email(mail):
            continue

        emails.append(mail)
        return emails


def extract_mail_reg(soup):
    text = soup.get_text()
    email_regex = re.compile(r'''
    ^([a-z0-9]+(?:[.-]?[a-z0-9]+)*    # username
    @
    [a-z0-9]+(?:[.-]?[a-z0-9]+)*      # domain
    \.[a-z]{2,7})$                    # domain 
    ''', re.VERBOSE | re.IGNORECASE)
    match = email_regex.findall(text)

    if len(match) == 0:
        match = re.findall(r'[\w\.-]+@[\w\.-]+', text)

    match = [e for e in match if is_valid_email(e)]
    match = [re.sub(r"^\d+", "", e) for e in []]

    return match


def find_email(prof, nth_page):
    possible_emails = []
    # scrape for the first 2 websites / 2 seconds pause
    for url in search(prof, tld="com", lang="en", num=1, start=nth_page, stop=1, pause=2):

        print(url)
        #addTk(frame3, url)

        if "researchgate" in url:
            print("research gate")
            continue
        # access to the url, move to the next web if any errors
        try:
            page = requests.get(url, timeout=10)
        except:
            continue

        # exclude pdf file
        content_type = page.headers.get("content-type")
        if "application/pdf" in content_type:
            print("pdf file")
            continue

        soup = BeautifulSoup(page.content, "html.parser")

        # find email by mailto html
        try:
            print("mailto")
            emails = extract_mailto(soup)
            possible_emails.extend(emails)
            if len(possible_emails) == 0:
                raise Exception("Not Found")
        except:
            print("re")
            # find email using regular expression
            match = extract_mail_reg(soup)
            possible_emails.extend(match)

    print(possible_emails)
    add_list_box(list_box_2, str(possible_emails))

    return possible_emails


def main(excel_file, start, end, nth_page, name_col, society_col, email_col):
    count = 0
    error = False
    # display text while downloading the search result page
    print('Searching...')
    
    # get prof name + org from excel file
    workbook = openpyxl.load_workbook(filename=excel_file)
    sheet = workbook[workbook.sheetnames[0]]

    for i in range(start, end):
        if sheet.cell(row=i, column=email_col).value != None:
            print(f"Found")
            continue

        name = sheet.cell(row=i, column=name_col).value
        uni = sheet.cell(row=i, column=society_col).value
        prof = name + " " + uni + " email"

        print()
        print(i, prof)
        info = f"{i}, {name}"
        add_list_box(list_box_2, "")
        add_list_box(list_box_2, info)

        try:
            possible_emails = find_email(prof, nth_page)
        except Exception as e:
            print(e)
            if str(e) == "HTTP Error 429: Too Many Requests":
                error = True
                break
            continue
        
        count += 1
        # update excel file
        try:
            print(possible_emails)
            sheet.cell(row=i, column=email_col).value = ' '.join(possible_emails)
        except TypeError as e:
            print(e)
            continue
        

        # update GUI
        root.update()

    # display time for app history
    time = datetime.now().strftime('%Y-%m-%d %H:%M')
    warn1 = f"{time}     Warning: Too Many Requests"
    warn2 = f"{time}     Success: Found from {start} to {i-1}"
    success1 = f"{time}     Success: Found from {start} to {i}"
    success2 = f"{time}     Success: Found {start}"

    if error:
        add_list_box(list_box, warn1)
        history_run.append(warn1)
        if count != 0:
            add_list_box(list_box, warn2)
            history_run.append(warn2)
    else:
        if count != 0:
            add_list_box(list_box, success1)
            history_run.append(success1)
        else:
            add_list_box(list_box, success2)
            history_run.append(success2)

    workbook.save(excel_file)
    print(f"Done!")


########################################3

# GUI
apps = []
if os.path.isfile("save.txt"):
    with open("save.txt", "r") as f:
        tempApps = f.read()
        tempApps = tempApps.split(",")
        apps = [x for x in tempApps if x.strip()]


root = tk.Tk()
root.title("Future Finder")
root.columnconfigure(0, minsize=250, weight=1)
root.rowconfigure([0, 1], minsize=100)
root.geometry("800x600")

history_run = []

def addFile():
    filename = filedialog.askopenfilename(initialdir="/", title="Select File", 
                                      filetypes=[("Excel", "*.xlsx")])
    print(filename)
    ent_dir.delete(0, "end")
    ent_dir.insert(0, filename)

def validate_input():
    try:
        inputs = [int(ent_from.get()), int(ent_to.get()), int(ent_nth.get()), int(ent_name.get()), int(ent_inst.get()), int(ent_email.get())]
        if ent_dir.get() == "" or int(ent_from.get()) >= int(ent_to.get())+1 or any(e < 1 for e in inputs):
            return False
        return True
    except ValueError:
        return False

def check_excel_open():
    try:
        with open(ent_dir.get(), "r+"):
            return True
    except IOError:
        return False

def runApps():
    runButton["state"] = "disabled"

    if validate_input() and check_excel_open():
        warn_input.config(text="")
        app = ent_dir.get().split("/")[-1]
        print(app)
        run = app + " From " + ent_from.get() + " To " + ent_to.get() + ' Page Rank ' + ent_nth.get()
        print(run)
        add_list_box(list_box, run)
        history_run.append(run)
        add_list_box(list_box_2, "Searching...")
        main(ent_dir.get(), int(ent_from.get()), int(ent_to.get())+1, int(ent_nth.get())-1, int(ent_name.get()), int(ent_inst.get()), int(ent_email.get()))
        add_list_box(list_box, " ")
        add_list_box(list_box_2, "---------------Done---------------")

        # Show message when the program is done
        messagebox.showinfo("Future Finder", "The program is done")
    else:
        warn_input.config(text="Invalid Input / Close Excel File")
    
    runButton["state"] = "normal"


def add_list_box(box, text):
    box.insert("end", text)

# Frame for input and button
frame1 = tk.Frame(root, bg="white")
frame1.place(relwidth=1, relheight=0.2)

# open file button
openFile = tk.Button(frame1, text="Open File", padx=10, 
                     pady=5, fg="white", bg="#263D42", command=addFile)
frm_entry = tk.Frame(frame1)
ent_dir = tk.Entry(master=frm_entry, width=50)

openFile.grid(row=0, column=0, pady=10)
frm_entry.grid(row=0, column=1, padx=10)
ent_dir.grid(row=0, column=0, sticky="e")

# display warning invalid input
warn_input = tk.Label(frame1, bg='#fff', fg='#f00')
warn_input.grid(row=0, column=2)

# scientist, inst, email column input
name_entry = tk.Frame(frame1)
ent_name = tk.Entry(master=name_entry, width=10)
lbl_name = tk.Label(master=name_entry, text="Scientist Col")

name_entry.grid(row=1, column=0, padx=5)
lbl_name.grid(row=0, column=0)
ent_name.grid(row=0, column=1)

inst_entry = tk.Frame(frame1)
ent_inst = tk.Entry(master=inst_entry, width=10)
lbl_inst = tk.Label(master=inst_entry, text="Institution Col")

inst_entry.grid(row=1, column=1, padx=5)
lbl_inst.grid(row=0, column=0, sticky="w")
ent_inst.grid(row=0, column=1, sticky="e")

email_entry = tk.Frame(frame1)
ent_email = tk.Entry(master=email_entry, width=10)
lbl_email = tk.Label(master=email_entry, text="Email Col")

email_entry.grid(row=1, column=2, padx=5)
lbl_email.grid(row=0, column=0, sticky="w")
ent_email.grid(row=0, column=1, sticky="e")

# from, to, page rank inout
from_entry = tk.Frame(frame1)
ent_from = tk.Entry(master=from_entry, width=10)
lbl_from = tk.Label(master=from_entry, text="From")

from_entry.grid(row=2, column=0, pady=5, padx=5)
lbl_from.grid(row=0, column=0, sticky="w")
ent_from.grid(row=0, column=1, sticky="e")

to_entry = tk.Frame(frame1)
ent_to = tk.Entry(master=to_entry, width=10)
lbl_to = tk.Label(master=to_entry, text="To")

to_entry.grid(row=2, column=1, pady=5, padx=5)
lbl_to.grid(row=0, column=0, sticky="w")
ent_to.grid(row=0, column=1, sticky="e")

nth_entry = tk.Frame(frame1)
ent_nth = tk.Entry(master=nth_entry, width=10)
lbl_nth = tk.Label(master=nth_entry, text="Page Rank")

nth_entry.grid(row=2, column=2, pady=5, padx=5)
lbl_nth.grid(row=0, column=0, sticky="w")
ent_nth.grid(row=0, column=1, sticky="e")


# Add Run button

runButton = tk.Button(frame1, text="Run", padx=20, 
                     pady=5, fg="white", bg="#263D42", command=lambda: threading.Thread(target=runApps).start())
runButton.grid(row=2, column=3, pady=10, padx=20)


# Add scrollbar 

frame2 =  tk.Frame(root)
frame2.place(relwidth=0.5, relheight=0.8, rely=0.2)

scrollbar = tk.Scrollbar(frame2)
scrollbar.pack(side="right", fill="y")
list_box = tk.Listbox(frame2, yscrollcommand= scrollbar.set)
list_box.pack(fill="both", expand=True)
scrollbar.config(command=list_box.yview)

frame3 = tk.Frame(root)
frame3.place(relwidth=0.5, relheight=0.8, relx=0.5, rely=0.2)

scrollbar_2 = tk.Scrollbar(frame3)
scrollbar_2.pack(side="right", fill="y")
list_box_2 = tk.Listbox(frame3, yscrollcommand= scrollbar_2.set)
list_box_2.pack(fill="both", expand=True)
scrollbar_2.config(command=list_box_2.yview)


for app in apps:
    list_box.insert("end", app)

root.mainloop()


with open("save.txt", "w") as f:
    for run in history_run:
        f.write(run + ",")